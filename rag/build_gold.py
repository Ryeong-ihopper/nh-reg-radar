# -*- coding: utf-8 -*-
"""심의 체크리스트·사례 → 검색 평가셋(gold).

**직접 만들지 않는다.** 「심의사례 NH농협은행 준법감시부 1.xlsx」의 체크리스트 시트는
준법감시부가 실제로 쓰는 표라, 점검항목마다 근거규정이 이미 붙어 있다.

  DEP-INC-06 | 예금성 | 포함사항 | 이자율·수익률의 범위 및 산출기준을 표시하였는가?
             | 기준 §16① 5 나

점검내용이 곧 질의, 근거규정이 곧 정답이다. 내가 광고를 읽고 정답을 지어내면 그건
내 판단을 재는 것이지 시스템을 재는 것이 아니다 — 사람이 만든 공식 매핑을 쓴다.

**약칭을 실제 규정명·조문키로 옮기는 것이 이 모듈의 일이다.** 「기준 §16① 5 나」는
사람에게는 자명하지만 검색 결과와 대조하려면 `은행 광고심의 기준 및 세칙 / 제16조`
로 바꿔야 한다. 항·호·목(① 5 나)까지는 청크 단위가 조문이라 버린다 — 조문 하나가
맞으면 맞은 것으로 본다.

  python rag/build_gold.py
  python rag/build_gold.py --check      # 정답 조문이 코퍼스에 실재하는지만 확인
"""
import os
import re
import sys
import json
import argparse
import collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = r"C:\Users\babie\OneDrive\Desktop\씨지인사이드\심의사례 NH농협은행 준법감시부 1.xlsx"
UNIFIED = (r"C:\Users\babie\OneDrive\Desktop\씨지인사이드"
           r"\(참고용) 데이터셋 작업 과정\01_핵심데이터"
           r"\중요_NH_광고심의_통합체크리스트.xlsx")
CHUNKS = os.path.join(ROOT, "output", "_rag", "chunks.jsonl")
RULE_INDEX = os.path.join(ROOT, "output", "_rag", "rule_index.jsonl")
OUT = os.path.join(ROOT, "output", "_rag", "gold.json")

# 약칭 → 코퍼스의 실제 규정명. 표에 쓰인 표기를 그대로 키로 둔다.
# 긴 것부터 맞춰야 「금소법 시행령」이 「금소법」으로 잘못 잡히지 않는다(아래 정렬).
ALIAS = {
    "기준": "은행 광고심의 기준 및 세칙",
    "세칙": "은행 광고심의 기준 및 세칙",
    "금소법 시행령": "금융소비자 보호에 관한 법률 시행령",
    "금소법": "금융소비자 보호에 관한 법률",
    "감독규정": "금융소비자 보호에 관한 감독규정",
    "정보통신망법 시행령": "정보통신망 이용촉진 및 정보보호 등에 관한 법률 시행령",
    "정보통신망법": "정보통신망 이용촉진 및 정보보호 등에 관한 법률",
    "금융지주회사법": "금융지주회사법",
    "광고에 관한 심사지침": "금융상품 등의 표시·광고에 관한 심사지침",
    "광고 심사지침": "금융상품 등의 표시·광고에 관한 심사지침",
    "표시·광고 심사지침": "금융상품 등의 표시·광고에 관한 심사지침",
    # ── 통합체크리스트(M04+M12)가 쓰는 약칭. 위와 표기 습관이 다르다 ──────
    # 「기준」 대신 「은행 광고심의 기준」, 「협회규정」처럼 업계 통용 줄임말을 쓴다.
    "협회규정": "금융투자회사의 영업 및 업무에 관한 규정",
    "은행 광고심의 기준": "은행 광고심의 기준 및 세칙",
    "금소법 감독규정": "금융소비자 보호에 관한 감독규정",
    "금융소비자보호감독규정": "금융소비자 보호에 관한 감독규정",
    "금융지주회사감독규정": "금융지주회사감독규정",
    "증발공규정": "증권의 발행 및 공시 등에 관한 규정",
    "금투업규정": "금융투자업규정",
    # 코퍼스 쪽 이름에 가운뎃점이 ㆍ(U+318D)라 · 로 적으면 못 찾는다. 실측으로 확인.
    "추천·보증 등에 관한 표시광고 심사지침": "추천ㆍ보증 등에 관한 표시ㆍ광고 심사지침",
    "추천·보증 등에 관한 표시·광고 심사지침": "추천ㆍ보증 등에 관한 표시ㆍ광고 심사지침",
    "금융상품 등의 표시·광고에 관한 심사지침": "금융상품 등의 표시·광고에 관한 심사지침",
    "예금자보호법": "예금자보호법",
    "인공지능기본법": "인공지능 발전과 신뢰 기반 조성 등에 관한 기본법",
    "약관법": "약관의 규제에 관한 법률",
    "퇴직연금감독규정": "퇴직연금감독규정",
}
_ALIAS_ORDER = sorted(ALIAS, key=len, reverse=True)

# §16① 5 나 · §22③ 4 · §50④ · §22조④ 3 나 · §2-40  — 조 번호만 뽑는다.
# 금투협 규정은 「§2-40」처럼 편-조 두 자리다. 앞자리를 떼면 제40조가 되어 엉뚱한
# 조문이 정답으로 박히므로 하이픈까지 잡는다.
_SEC = re.compile(r"§\s*(\d+(?:\s*-\s*\d+)?)\s*(?:조)?\s*(?:의\s*(\d+))?")
# 「제9조제3항」처럼 § 없이 적힌 것도 있다(통합체크리스트의 단서 문구 안).
_ART = re.compile(r"제\s*(\d+(?:-\d+)?)\s*조(?:\s*의\s*(\d+))?")
# 감독규정 별표 3 같은 별표 참조
_TBL = re.compile(r"별표\s*0*(\d+)")


# 조문 표기(§ · 제N조 · 별표) 앞에 붙은 이름 부분
_NAME = re.compile(r"^(.*?)(?=§|별표|제\s*\d+\s*조|$)")
# 「… 심사지침 Ⅴ. 5」의 목차 기호. 이름 뒤 공백에 붙은 것만 본다.
_ROMAN = re.compile(r"\s([ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+)\s*[.．]")
# 사람이 로마숫자를 대문자 V·I·X 로 적기도 한다(실측: 「심사지침 V. 5」).
_LATIN2ROMAN = str.maketrans({"I": "Ⅰ", "V": "Ⅴ", "X": "Ⅹ"})


def _unknown_name(part):
    """ALIAS 에 없는 규정명이 앞에 붙어 있으면 그 이름을, 아니면 None.

    「§16① 2」처럼 이름 없이 조문만 있는 조각은 앞의 것을 물려받아야 하므로
    None 을 준다. 이름이 있는데 못 알아보는 것만 골라낸다.
    """
    nm = _NAME.match(part).group(1).strip(" ·,")
    return nm if len(nm) >= 3 else None


def parse_ref(text):
    """근거규정 문자열 → [(규정명, 조문키), ...]

    「금소법 §22③ 4, 금소법 시행령 §18① 1 다」처럼 여럿이 콤마로 붙는다.
    쉼표로 자른 뒤 조각마다 규정명을 찾되, 규정명이 생략된 조각은 **앞 조각의 것을
    물려받는다**(「기준 §16① 2, 기준 §16① 8」은 둘 다 있지만 「금소법 §22⑦,
    시행령 §19①, 감독규정 §18」처럼 섞이는 경우가 있어 순서대로 흘려보낸다).
    """
    out, cur = [], None
    # 엑셀 셀에 줄바꿈 없는 공백(\xa0)이 섞여 있다. 눈에는 보통 공백과 똑같아서
    # 「추천·보증 등에 관한 관한\xa0표시…」가 약칭과 안 맞는 걸 알아채기 어렵다.
    text = str(text or "").replace("\xa0", " ")
    # 통합체크리스트는 근거 여럿을 **줄바꿈**으로 쌓아 둔다(「§16① 2\n§16① 8」).
    # 쉼표만 자르면 둘째 줄이 첫 줄에 붙어 통째로 버려진다.
    for part in re.split(r"[,\n]", text):
        part = part.strip()
        if not part:
            continue
        matched = None
        for a in _ALIAS_ORDER:
            if part.startswith(a) or f" {a} " in f" {part} ":
                matched = ALIAS[a]
                break
        if matched:
            cur = matched
        elif part.startswith("시행령") and cur:
            # 「정보통신망법 §50④, 시행령 §61③」의 시행령은 **바로 앞 법**의 시행령이다.
            # 이걸 한 법으로 고정해 두면 조용히 틀린 정답이 박힌다(실측: 금소법
            # 시행령 제61조로 잘못 잡혀 3건이 코퍼스에서 안 나왔다 — 안 나와서
            # 들켰지, 우연히 존재하는 조문이었으면 못 잡을 뻔했다).
            cur = cur if cur.endswith("시행령") else cur + " 시행령"
        elif _unknown_name(part):
            # **모르는 규정명은 물려받게 두면 안 된다.** 「금융지주회사법 §48④,
            # 금융지주회사감독규정 §24」에서 뒤 조각이 앞의 법을 물려받으면 있지도
            # 않은 「금융지주회사법 제24조」가 정답으로 박히고 아무도 모른다.
            # 못 읽는 이름으로 남겨 두면 아래 미해결 보고에 걸린다.
            out.append((f"?{_unknown_name(part)}", ""))
            cur = None
            continue
        if not cur:
            continue
        # 심사지침류는 조문이 없고 「Ⅴ. 1. 가」 같은 목차로 근거를 적는다.
        # 코퍼스 키도 목차 기호(Ⅴ)라 대문자 V 로 적힌 것을 로마숫자로 맞춰 준다.
        rom = _ROMAN.search(part.translate(_LATIN2ROMAN))
        if rom and not _SEC.search(part):
            out.append((cur, rom.group(1)))
            continue
        hits = list(_SEC.finditer(part)) or list(_ART.finditer(part))
        for m in hits:
            n = re.sub(r"\s+", "", m.group(1))
            sub = m.group(2)
            out.append((cur, f"제{n}조" + (f"의{sub}" if sub else "")))
        for m in _TBL.finditer(part):
            out.append((cur, f"[별표 {int(m.group(1)):04d}]"))
    # 중복 제거(순서 유지)
    seen, uniq = set(), []
    for x in out:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


def load_chunks():
    rows = [json.loads(l) for l in open(CHUNKS, encoding="utf-8")]
    by = collections.defaultdict(list)
    for i, r in enumerate(rows):
        by[(r["reg"], r.get("key", ""))].append(i)
    return rows, by


def resolve(by, reg, key):
    """(규정명, 조문키) → 청크 번호 목록. 정확히 없으면 접두 일치로 한 번 더 본다."""
    hit = by.get((reg, key))
    if hit:
        return hit
    # 「제16조」로 찾는데 코퍼스 키가 「제16조의2」뿐인 경우 등
    out = []
    for (r, k), v in by.items():
        if r == reg and k.startswith(key):
            out += v
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--out", default=OUT)
    ap.add_argument("--check", action="store_true")
    a = ap.parse_args()

    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl

    wb = openpyxl.load_workbook(a.xlsx, read_only=True, data_only=True)
    rows, by = load_chunks()

    gold, unresolved = [], []

    # ── 1) 체크리스트 97항목 — 공식 매핑이라 가장 신뢰도가 높다 ────────────
    for r in wb["체크리스트"].iter_rows(values_only=True):
        if not r or not r[0] or str(r[0]).count("-") != 2:
            continue
        item_id, ptype, kind, q, ref = (str(x or "").strip() for x in r[:5])
        refs = parse_ref(ref)
        idxs, miss = [], []
        for reg, key in refs:
            got = resolve(by, reg, key)
            (idxs.extend(got) if got else miss.append(f"{reg} {key}"))
        if miss:
            unresolved.append((item_id, ref, miss))
        if not idxs:
            continue
        gold.append({"id": item_id, "출처": "체크리스트", "상품유형": ptype,
                     "구분": kind, "q": q, "근거원문": ref,
                     "정답": [{"reg": reg, "key": key} for reg, key in refs],
                     "정답청크": sorted(set(idxs))})

    # ── 2) 심의사례 20건 — 실제로 지적된 건이라 질의가 더 현실적이다 ───────
    for sheet, ptype in (("예금성 광고", "예금성"), ("대출성 광고", "대출성")):
        for r in wb[sheet].iter_rows(min_row=4, values_only=True):
            if not r or r[0] in (None, "") or not str(r[0]).strip().isdigit():
                continue
            no, name, _, _, _, _, chk, ref, op, detail = (
                [str(x or "").strip() for x in r[:10]] + [""] * 10)[:10]
            refs = parse_ref(ref)
            idxs, miss = [], []
            for reg, key in refs:
                got = resolve(by, reg, key)
                (idxs.extend(got) if got else miss.append(f"{reg} {key}"))
            if miss:
                unresolved.append((f"{ptype}#{no}", ref, miss))
            if not idxs:
                continue
            # 질의는 '점검내용 + 실제 지적사유'로 만든다 — 체크리스트만 쓰면
            # 일반론이 되고, 지적사유를 붙여야 그 광고의 상황이 들어간다.
            q = chk if not detail else f"{chk} ({op}: {detail})"
            gold.append({"id": f"{ptype}-사례{no}", "출처": "심의사례",
                         "상품유형": ptype, "구분": op, "q": q, "광고명": name,
                         "근거원문": ref,
                         "정답": [{"reg": reg, "key": key} for reg, key in refs],
                         "정답청크": sorted(set(idxs))})

    # ── 3) 통합체크리스트 339문항 — M04+M12 를 사람이 합친 표 ─────────────
    # 위 두 시트보다 규모가 3배고 조·항·호까지 근거가 붙어 있다. 특히 **투자성
    # 162문항**은 위 두 시트에 아예 없던 영역이라, 이걸 빼면 투자성 검색 성능을
    # 한 번도 재지 않은 채 넘어가게 된다.
    if os.path.exists(UNIFIED):
        wb2 = openpyxl.load_workbook(UNIFIED, read_only=True, data_only=True)
        for r in wb2["전체_체크리스트"].iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            no, src, page, ptype, big, mid, ref, orig, q = (
                [str(x or "").strip() for x in r[:9]] + [""] * 9)[:9]
            refs = parse_ref(ref)
            if not refs:
                continue                      # 근거 없는 117문항은 정답이 없다
            idxs, miss = [], []
            for reg, key in refs:
                got = resolve(by, reg, key)
                (idxs.extend(got) if got else miss.append(f"{reg} {key}"))
            if miss:
                unresolved.append((f"통합#{no}", ref.replace("\n", " / "), miss))
            if not idxs:
                continue
            gold.append({"id": f"U-{int(no):03d}", "출처": "통합체크리스트",
                         "상품유형": ptype, "구분": big, "q": q or orig,
                         "출처매뉴얼": src, "페이지": page,
                         "근거원문": ref.replace("\n", " / "),
                         "정답": [{"reg": rg, "key": k} for rg, k in refs],
                         "정답청크": sorted(set(idxs))})
        wb2.close()

    # ── 규칙을 정답으로 치려다 되돌린 자리 ────────────────────────────────
    # 「같은 조문을 근거로 든 규칙은 다 정답」으로 붙여 봤다가 뺐다. 조문 하나에
    # 규칙이 너무 많이 달려서 채점이 무의미해진다.
    #
    #   금투협 규정 제2-38조 → 규칙 161개    금소법 제22조 → 103개
    #   문항당 정답규칙 중앙 48개 · 최대 316개
    #
    # 8,309개 중 161개가 정답이면 상위 5개에 드는 것은 거의 자동이다. 실제로
    # 「은행의 명칭을 표시하였는가?」에 「예금성 의무표시 - 이자율의 범위 및
    # 산출기준」이 정답 처리됐다 — 둘 다 은행 광고심의 기준 제16조를 근거로
    # 든다는 이유뿐이고, 제16조 안의 항·호는 서로 다르다.
    #
    # 항·호까지 맞추면 되지만 gold 의 60%, 규칙 근거의 57% 에만 항·호가 있다.
    # 기준이 섞인 숫자는 읽을 수가 없다. **규칙 단위 정답표는 사람이 붙여야 한다.**

    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(gold, f, ensure_ascii=False, indent=1)

    print(f"gold {len(gold)}건 · 출처 "
          f"{collections.Counter(g['출처'] for g in gold).most_common()}")
    print(f"상품유형 {collections.Counter(g['상품유형'] for g in gold).most_common()}")
    n = [len(g['정답청크']) for g in gold]
    print(f"항목당 정답청크: 중앙 {sorted(n)[len(n)//2]} · 최소 {min(n)} · 최대 {max(n)}")
    print(f"저장: {a.out}")

    if unresolved:
        print(f"\n⚠ 코퍼스에서 못 찾은 참조 {len(unresolved)}건 — "
              f"약칭 매핑이나 조문 표기를 확인해야 한다")
        for i, (iid, ref, miss) in enumerate(unresolved[:12]):
            print(f"  {iid:16s} {ref[:40]:40s} → {', '.join(miss)}")


if __name__ == "__main__":
    main()
